from openpyxl.workbook import Workbook
from openpyxl.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import colors


class ExcelHandler:
    def __init__(self, filename = None, workbook = None):
        if filename is None:
            if isinstance(workbook, Workbook):
                self.__workbook__ = workbook
                self.__filename__ = "Default.xlsx"
            else:
                raise TypeError("Invalid type of workbook")
        else:
            self.open(filename)
            self.__filename__ = filename
        if self.__workbook__ is not None:
            self.__prepared__ = True
        else:
            self.__prepared__ = False

    @property
    def filename(self):
        return self.__filename__

    @filename.setter
    def filename(self, filename):
        self.__filename__ = filename
        self.open(filename)

    @property
    def workbook(self):
        return self.__workbook__

    @workbook.setter
    def workbook(self, workbook):
        if isinstance(workbook, Workbook):
            self.__workbook__ = workbook
        else:
            raise TypeError("Invalid argument of workbook {}".format(workbook))

    def open(self, filename):
        filename = str(filename)
        if filename.endswith("xls") or filename.endswith("xlsx"):
            self.workbook = load_workbook(filename)
            if self.workbook is not None:
                self.__prepared__ = True
        else:
            raise ValueError("Unsupported file format - {}".format(filename))

    def choose_sheet(self, sheetname):
        if self.__prepared__:
            return self.workbook.get_sheet_by_name(sheetname)
        else:
            raise RuntimeError("Not ready yet, please open excel first!")

    def edit(self, row, col, value = None, sheet = None, font = Font()):
        active_sheet = None
        if isinstance(sheet, str):
            for worksheet in self.workbook.worksheets:
                if worksheet.title == sheet:
                    active_sheet = worksheet
                    break
            if active_sheet is None:
                active_sheet = self.workbook.create_sheet(sheet)
        elif isinstance(sheet, Worksheet):
            active_sheet = sheet
        elif sheet is None:
            active_sheet = self.workbook.active
        active_sheet.cell(column=col, row=row).font = font
        active_sheet.cell(column=col, row=row).value = value if value is not None else ""

    def read(self, row = None, col = None, sheet = None):
        content = dict()

        def remove(sheet, col, row):
            s = list()
            if sheet is not None:
                if isinstance(sheet, Worksheet):
                    for name in self.workbook.get_sheet_names():
                        if sheet == self.workbook.get_sheet_by_name(name):
                            sheet = name
                            break
                for c in content:
                    if c != sheet:
                        s.append(c)
                for n in s:
                    del content[n]

            s.clear()
            if col is not None:
                for k in content:
                    for key in content[k]:
                        if key[0] != col:
                            s.append((k, key))
            for i, j in s:
                del content[i][j]

            s.clear()
            if row is not None:
                for k in content:
                    for key in content[k]:
                        if key[1] != row:
                            s.append((k, key))
            for i, j in s:
                del content[i][j]

        for name in self.workbook.get_sheet_names():
            ws = self.workbook.get_sheet_by_name(name)
            cell_content = dict()
            for cols in ws.columns:
                for c in cols:
                    if c.value is not None:
                        cell_content[(c.column, c.row)] = c.value
            content[name] = cell_content

        remove(sheet, col, row)

        return  content

    def close(self, filename = None):
        self.__workbook__.save(filename if isinstance(filename, str) else self.__filename__)

if __name__ == '__main__':
    excel = ExcelHandler(filename="Empty.xlsx")
    excel.edit(1, 2, 'B2')
    excel.edit(3, 4, "测试数据")
    font = Font(name="微软雅黑", size=15, color=colors.RED)
    for c in excel.workbook.active.columns:
        print(c)
    excel.edit(4, 5, "红色的", font = font)
    print(excel.workbook.active.rows)
    print(excel.workbook.active.columns)
    print(excel.workbook.get_sheet_names())
    print(excel.read(sheet="Sheet1", col='E'))
    excel.close()