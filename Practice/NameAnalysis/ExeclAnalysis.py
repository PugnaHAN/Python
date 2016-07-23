from openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.styles import Font
from NameAnalysis.NameDescriptor import NameDescriptor
from openpyxl.cell import Cell


class ExeclDescriptor:
    def __init__(self, filename):
        self.__work_book__ = load_workbook(filename)
        self.__sheets__ = self.__work_book__.get_sheet_names()
        self.__file_name__ = filename

    @property
    def work_book(self):
        return self.__work_book__

    @property
    def sheets(self):
        return self.__sheets__

    @property
    def file_name(self):
        return self.__file_name__

__file_name__ = "Test.xlsx"
excel = ExeclDescriptor(__file_name__)

def read_excel(sheetname):
    sheet = excel.work_book.get_sheet_by_name(sheetname)
    # read the cols and rows
    nrows = len(sheet.rows)
    # Initialize filenames
    filenames = list()
    for i in range(nrows):
        filenames.append(sheet.cell(row = i + 1, column=1).value)
    # print(filenames)
    return filenames

sheet_created = False

def write_excel(sheet, col, row, value, font):
    global sheet_created
    if sheet_created == False:
        work_sheet = excel.work_book.create_sheet(sheet)
    else:
        work_sheet = excel.work_book.get_sheet_by_name(sheet)
    sheet_created = True
    work_sheet.font = font
    work_sheet.cell(row = row, column= col, value = value)

def save():
    excel_writer = ExcelWriter(excel.work_book)
    excel_writer.save(filename=excel.file_name)

def set_font(font_name = "Times New Roman", size = 12, color = 'FF000000', bold = False):
    font = Font(name= font_name, size=size, bold=bold, color=color)
    return font

def write_data(name_descriptor, row):
    if isinstance(name_descriptor, NameDescriptor):
        data = list()
        data.append(name_descriptor.src_name)
        data.append(str(name_descriptor.paper_size))
        data.append(name_descriptor.paper_type)
        data.append(str(name_descriptor.print_mode))
        data.append(name_descriptor.full_bleed)
        if(name_descriptor.src_name == 'Unknown'):
            font = set_font(color='FFFF0000')
        else:
            font = set_font()
        for i in range(len(data)):
            write_excel(sheet='files', col=i + 1, row=row, font=font, value=data[i])
    else:
        raise TypeError

def init_header():
    font = set_font(size=14, bold=True)
    write_excel('files', 1, 1, 'File Name', font)
    write_excel('files', 2, 1, 'Paper Size', font)
    write_excel('files', 3, 1, 'Paper Type', font)
    write_excel('files', 4, 1, 'Print Mode', font)
    write_excel('files', 5, 1, 'Full Bleed', font)

def main():
    # Read the excel file
    origin_data = read_excel("life printing")
    #Initialzie the header
    init_header()
    #Write the remaining data into sheet
    for i in range(len(origin_data)):
        name_descriptor = NameDescriptor()
        print(origin_data[i])
        name_descriptor.analysis_from_filename(origin_data[i])
        write_data(name_descriptor, i + 2)
    # save file
    save()

if __name__ == '__main__':
    main()